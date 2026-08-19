# -*- coding: utf-8 -*-
"""Phase 1a: 매체별 RAW 파일 읽기 → 정규화 → 광고비 보정 → 매칭키 산출.

각 리더는 표준 스키마 DataFrame 반환:
  매체, 브랜드, 캠페인, 광고그룹, 광고(소재), 날짜(Timestamp),
  광고비_raw, 노출수, 클릭수, 매칭키
광고비용(보정) = 광고비_raw * COST_COEF[매체] 는 combine_ads()에서 적용.
"""
import re
import calendar
import warnings
warnings.simplefilter("ignore")   # openpyxl 기본스타일 경고 등 숨김
import pandas as pd
from openpyxl import load_workbook
from config import RAW_DIR, COST_COEF, BRANDS, JEONGAEK

STD = ["매체", "브랜드", "캠페인", "광고그룹", "광고(소재)", "날짜",
       "광고비_raw", "노출수", "클릭수", "매칭키"]


# ---------- 유틸 ----------
def to_num(x):
    if x is None:
        return 0.0
    if isinstance(x, (int, float)):
        return float(x)
    s = str(x).replace("\xa0", "").replace('"', "").replace(",", "").replace(" ", "").strip()
    if s == "" or s.lower() == "nan":
        return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0


def to_date(x):
    if isinstance(x, pd.Timestamp):
        return x.normalize()
    s = str(x).strip().rstrip(".")
    s = s.replace(".", "-").replace("/", "-")
    ts = pd.to_datetime(s, errors="coerce")
    return ts.normalize() if pd.notna(ts) else pd.NaT


def brand_from(campaign, idx=1):
    toks = re.split(r"[_/ ]", str(campaign))
    if len(toks) > idx and toks[idx] in BRANDS:
        return toks[idx]
    for t in toks:
        if t in BRANDS:
            return t
    return ""


def norm_id(x):
    """광고/캠페인 ID → Excel 15자리 유효숫자 정규화 문자열. Meta 매칭용."""
    v = to_num(x)
    if v <= 0:
        return ""
    return "%.15g" % v


def _code(text, prefix):
    """text에서 첫 'prefix\\d+' 코드 추출 (KK/NG/CT). 없으면 ''"""
    m = re.search(prefix + r"\d+", str(text))
    return m.group(0) if m else ""


def _xlsx_rows(path, sheet=None):
    wb = load_workbook(path, read_only=False, data_only=True)  # Criteo read_only 버그 회피
    ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    return rows


def _read_csv(path, encoding, sep=",", skiprows=0):
    return pd.read_csv(path, encoding=encoding, sep=sep, skiprows=skiprows,
                       dtype=str, keep_default_na=False)


def _find(subdir, pattern):
    """RAW/subdir 안에서 pattern에 맞는 첫 파일 반환. 없으면 None (건너뜀 처리용)."""
    return next(iter(sorted((RAW_DIR / subdir).glob(pattern))), None)


# ---------- 매체별 리더 (매칭키 포함) ----------
def read_meta():
    out = []
    for f in sorted((RAW_DIR / "Meta").glob("*.xlsx")):
        for r in _xlsx_rows(f, "Raw Data Report")[1:]:
            if r[0] is None:
                continue
            camp = str(r[1] or ""); cre = str(r[3] or "")
            key = _code(cre, "MT") or camp                    # 매칭키 = 소재의 MT코드
            out.append(["Meta", brand_from(camp), camp, str(r[2] or ""), cre,
                        to_date(r[0]), to_num(r[5]), to_num(r[6]), to_num(r[7]), key])
    return pd.DataFrame(out, columns=STD)


def read_google():
    f = _find("Google", "*.csv")
    if f is None:
        return pd.DataFrame(columns=STD)
    df = _read_csv(f, "utf-16", sep="\t", skiprows=2)
    df.columns = [c.strip() for c in df.columns]
    out = []
    for _, r in df.iterrows():
        camp = str(r["캠페인"]).strip()
        if not camp:
            continue
        out.append(["Google", brand_from(camp), camp, camp, camp,
                    to_date(r["일"]), to_num(r["비용"]), to_num(r["노출수"]), to_num(r["클릭수"]),
                    camp])                                    # 매칭키 = 캠페인명
    return pd.DataFrame(out, columns=STD)


def read_kko():
    f = _find("KKO", "*.csv")
    if f is None:
        return pd.DataFrame(columns=STD)
    df = _read_csv(f, "utf-16", sep="\t")
    df.columns = [c.strip().strip('"') for c in df.columns]
    out = []
    for _, r in df.iterrows():
        camp = str(r["캠페인 이름"]).strip()
        if not camp:
            continue
        cre = str(r["소재 이름"]).strip()
        key = _code(cre, "KK") or camp
        out.append(["KKO", brand_from(camp), camp, str(r["광고그룹 이름"]).strip(), cre,
                    to_date(r["일"]), to_num(r["비용"]), to_num(r["노출수"]), to_num(r["클릭수"]),
                    key])
    return pd.DataFrame(out, columns=STD)


def read_criteo():
    f = _find("Criteo", "*.xlsx")
    if f is None:
        return pd.DataFrame(columns=STD)
    out = []
    for r in _xlsx_rows(f, "Download")[1:]:
        if r[0] is None:
            continue
        camp = str(r[1] or ""); cre = str(r[5] or "")
        typ = camp.split("_pf_")[-1] if "_pf_" in camp else ""
        key = f"{brand_from(camp)}_{typ}".upper()      # 브랜드+유형(캠페인) 단위 매칭
        out.append(["Criteo", brand_from(camp), camp, str(r[3] or ""), cre,
                    to_date(r[0]), to_num(r[6]), to_num(r[7]), to_num(r[8]), key])
    return pd.DataFrame(out, columns=STD)


def read_rtb():
    f = _find("RTB", "*.xlsx")
    if f is None:
        return pd.DataFrame(columns=STD)
    out = []
    for r in _xlsx_rows(f)[1:]:
        if r[0] is None:
            continue
        camp = str(r[0])                       # 'IT_pf'
        brand = brand_from(camp, idx=0)
        out.append(["RTB", brand, camp, camp, camp,
                    to_date(r[1]), to_num(r[5]), to_num(r[2]), to_num(r[3]),
                    brand])                                   # 매칭키 = 브랜드
    return pd.DataFrame(out, columns=STD)


def read_naver_advoost():
    f = _find("NAV", "*Advoost*.csv")
    if f is None:
        return pd.DataFrame(columns=STD)
    df = _read_csv(f, "utf-8-sig")
    df.columns = [c.strip() for c in df.columns]
    out = []
    for _, r in df.iterrows():
        camp = str(r["캠페인 이름"]).strip()
        if not camp:
            continue
        out.append(["Naver", brand_from(camp), camp, camp, camp, to_date(r["기간"]),
                    to_num(r["총비용"]), to_num(r["노출수"]), to_num(r["클릭수"]),
                    camp])                                    # advoost 매칭키 = 캠페인명
    return pd.DataFrame(out, columns=STD)


def read_naver_smart():
    f = _find("NAV", "*Smart*.csv")
    if f is None:
        return pd.DataFrame(columns=STD)
    df = _read_csv(f, "utf-8-sig")
    df.columns = [c.strip() for c in df.columns]
    out = []
    for _, r in df.iterrows():
        camp = str(r["캠페인 이름"]).strip()
        if not camp:
            continue
        cre = str(r["광고 소재 이름"]).strip()
        key = _code(cre, "NG") or camp
        out.append(["Naver", brand_from(camp), camp, str(r["광고 그룹 이름"]).strip(), cre,
                    to_date(r["기간"]), to_num(r["총비용"]), to_num(r["노출수"]), to_num(r["클릭수"]),
                    key])
    return pd.DataFrame(out, columns=STD)


def read_nsa():
    f = _find("NSA", "*.csv")
    if f is None:
        return pd.DataFrame(columns=STD)
    df = _read_csv(f, "utf-8-sig", skiprows=1)   # 1행 제목 skip
    df.columns = [c.strip() for c in df.columns]
    out = []
    for _, r in df.iterrows():
        camp = str(r["캠페인"]).strip()
        if not camp:
            continue
        out.append(["Naver SA", brand_from(camp), camp, camp, camp, to_date(r["일별"]),
                    to_num(r["총비용"]), to_num(r["노출수"]), to_num(r["클릭수"]),
                    camp])                                    # NSA 매칭키 = 캠페인(SA조인 별도)
    return pd.DataFrame(out, columns=STD)


def read_dable():
    f = _find("Dable", "*.xlsx")
    if f is None:
        return pd.DataFrame(columns=STD)
    out = []
    for r in _xlsx_rows(f, "ad-performance-stats")[1:]:
        if r[0] is None:
            continue
        camp = str(r[2] or ""); cre = str(r[5] or "")       # 캠페인명, 콘텐츠관리용제목(소재)
        key = _code(cre, "DB") or camp                       # 애드코드 DB####
        out.append(["Dable", brand_from(camp), camp, camp, cre,
                    to_date(r[0]), to_num(r[8]), to_num(r[6]), to_num(r[7]), key])
    return pd.DataFrame(out, columns=STD)                    # 소비/유효노출/클릭


def read_tiktok():
    f = _find("TikTok", "*.xlsx")
    if f is None:
        return pd.DataFrame(columns=STD)
    out = []
    for r in _xlsx_rows(f, "Sheet1")[1:]:
        camp = str(r[0] or "")
        if not camp or camp.startswith("총"):                 # '총 N개 결과' 요약행 skip
            continue
        cre = str(r[2] or "")                                # 광고 이름(소재)
        key = _code(cre, "TT") or camp                       # 애드코드 TT####
        out.append(["TikTok", brand_from(camp), camp, str(r[1] or ""), cre,
                    to_date(r[3]), to_num(r[4]), to_num(r[5]), to_num(r[6]), key])
    return pd.DataFrame(out, columns=STD)                    # 일별/지출/노출/클릭(목적지)


def read_toss():
    f = _find("Toss", "*.csv")
    if f is None:
        return pd.DataFrame(columns=STD)
    df = _read_csv(f, "utf-8-sig")
    df.columns = [c.strip() for c in df.columns]
    out = []
    for _, r in df.iterrows():
        camp = str(r["캠페인명"]).strip()                    # 끝 탭문자 제거
        if not camp:
            continue
        cre = str(r["소재명"]).strip()                       # 소재명
        key = _code(cre, "TS") or camp                       # 애드코드 TS####
        out.append(["Toss", brand_from(camp), camp, str(r["광고세트명"]).strip(), cre,
                    to_date(r["이벤트 발생 날짜"]), to_num(r["집행 비용 (VAT 제외) (₩)"]),
                    to_num(r["노출 수"]), to_num(r["클릭 수"]), key])
    return pd.DataFrame(out, columns=STD)


READERS = [read_meta, read_google, read_kko, read_criteo, read_rtb,
           read_naver_advoost, read_naver_smart, read_nsa,
           read_dable, read_tiktok, read_toss]


def build_jeongaek(year, month):
    """정액을 그달 1일~[과거월=말일 / 진행월=어제]까지 매일 budget/그달일수씩 배분.
    → 완료된 달=전액, 진행 중인 달=경과일수 비례(시간 기반 고정비, 광고데이터 유무 무관)."""
    days_in_month = calendar.monthrange(year, month)[1]
    start = pd.Timestamp(year, month, 1)
    month_end = pd.Timestamp(year, month, days_in_month)
    yesterday = pd.Timestamp.now().normalize() - pd.Timedelta(days=1)
    end = min(month_end, yesterday)                  # 진행 중인 달이면 어제까지만 누적
    out = []
    if end >= start:
        for camp, (brand, key, budget) in JEONGAEK.items():
            daily = budget / days_in_month
            for d in pd.date_range(start, end, freq="D"):
                out.append(["Naver SA", brand, camp, "정액", "정액", d.normalize(),
                            daily, 0.0, 0.0, key])
    return pd.DataFrame(out, columns=STD)


def combine_ads(target=None):
    """target=(연,월)이면 그 달 기준으로 정액 배분. 없으면 데이터 최소월 기준."""
    parts = [fn() for fn in READERS]
    df = pd.concat(parts, ignore_index=True)
    df["날짜"] = pd.to_datetime(df["날짜"], errors="coerce")
    df = df[df["날짜"].notna()].reset_index(drop=True)
    # 리포트 실행 당일(및 그 이후) 데이터 제외 — 당일치는 아직 미확정이므로 뺀다.
    today = pd.Timestamp.now().normalize()
    df = df[df["날짜"] < today].reset_index(drop=True)
    if df.empty:
        raise SystemExit("실행 당일 데이터를 제외하니 남은 데이터가 없습니다. (전일 이전 RAW가 있는지 확인하세요)")
    y, m = target if target else (df["날짜"].min().year, df["날짜"].min().month)
    jg = build_jeongaek(y, m)
    df = pd.concat([df, jg], ignore_index=True)
    df["날짜키"] = df["날짜"].dt.strftime("%Y%m%d")
    # 소재 단위 집계 (같은 소재가 여러 광고세트로 쪼개진 경우 합산 → GA 중복부여 방지)
    dims = ["매체", "브랜드", "캠페인", "광고그룹", "광고(소재)", "날짜", "날짜키", "매칭키"]
    df = (df.groupby(dims, as_index=False)
            .agg({"광고비_raw": "sum", "노출수": "sum", "클릭수": "sum"}))
    df["광고비용"] = df.apply(lambda r: r["광고비_raw"] * COST_COEF[r["매체"]], axis=1)
    return df


if __name__ == "__main__":
    df = combine_ads()
    g = df.groupby("매체")[["광고비용", "노출수", "클릭수"]].sum()
    print(g.to_string())
    print("\nrows:", len(df), " 광고비 합계:", f"{df['광고비용'].sum():,.0f}")
