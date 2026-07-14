# -*- coding: utf-8 -*-
"""Phase 5: 브랜드 종합 + 리포트 추가 요청.

- 브랜드 종합: 브랜드별 [매체 총 누적] 세로 스택.
- 리포트 추가 요청: 일별 광고비 매트릭스 (브랜드×매체세부 행 × 날짜 열 + 월누계).
"""
import warnings
warnings.simplefilter("ignore")
import calendar
from datetime import date
from openpyxl.styles import Alignment, Border, Side, PatternFill, Font
from total import (MEDIA_ROWS, _slice, _metrics, _put, F_TITLE, F_SEC, F_COL, F_SUM,
                   FILL_SEC, FILL_COL, FILL_SUM, CENTER, LEFT, BRAND_TITLE)

# Total [매체 예산 집행율]의 월예산(MEDIA_ROWS) → (매체, 패턴) 집행율 기준
# 패턴 앞 '_' 정규화(_ca→ca), 데이터원 없는 _NONE_(신규매체) 제외
BS_BUDGET = {(md, pt.lstrip("_")): bd
             for _g, _lb, md, pt, bd in MEDIA_ROWS if md != "_NONE_"}
# 소계 집행율 기준 = 브랜드 총예산 (Total 합계 예산과 동일: MEDIA_ROWS 월예산 총합)
TOTAL_BUDGET = sum(bd for *_, bd in MEDIA_ROWS)

# 브랜드 종합 지표 (참고파일 순서)
BS_KEYS = ["노출수", "클릭수", "클릭률", "클릭당비용", "집행예산", "전환수", "매출",
           "회원가입", "전환율", "전환당비용", "회원가입율", "ROAS", "객단가"]
BS_FMT = ["#,##0", "#,##0", "0.00%", "#,##0", "#,##0", "#,##0", "#,##0",
          "#,##0", "0.00%", "#,##0", "0.00%", "#,##0.00", "#,##0"]

# 브랜드 종합_F 매체 taxonomy: (유형, 매체라벨, 통합매체, 캠페인패턴)
BS_MEDIA = [
    ("성과형", "네이버 브랜드검색", "Naver SA", "bsa"),
    ("성과형", "네이버 키워드검색", "Naver SA", "cpc"),
    ("성과형", "네이버 쇼핑검색", "Naver SA", "shopping"),
    ("성과형", "네이버 플레이스", "Naver SA", "place"),
    ("성과형", "네이버 엠버서더형", "Naver SA", "Ambassador"),
    ("성과형", "구글 키워드검색", "Google", "cpc"),
    ("성과형", "네이버 성과형DA - 스마트채널", "Naver", "smart"),
    ("성과형", "네이버 성과형DA - ADVoost", "Naver", "advoost"),
    ("성과형", "카카오 네이티브", "KKO", "ntv"),
    ("성과형", "카카오 비즈보드", "KKO", "biz"),
    ("성과형", "카카오 카탈로그", "KKO", "ca"),
    ("성과형", "구글 쇼핑", "Google", "pmax"),
    ("성과형", "크리테오", "Criteo", ""),
    ("성과형", "크루비", "_NONE_", ""),
    ("성과형", "RTB하우스", "RTB", ""),
    ("성과형", "인스타그램 ", "Meta", "pf"),
    ("성과형", "구글 GDN", "Google", "gdn"),
    ("노출형", "인스타그램 ", "Meta", "br"),
    ("노출형", "구글 YouTube", "Google", "youtube"),
]
NAVY = PatternFill("solid", fgColor="002060")
F_BANNER = Font(name="맑은 고딕", bold=True, size=16, color="FFFFFF")
F_BANNER_BIG = Font(name="맑은 고딕", bold=True, size=24, color="FFFFFF")


def write_brand_summary(ws, uni, y, mth):
    ncol = 6 + len(BS_KEYS)          # 마지막 데이터 열(집행율=6, 지표 7~19)
    # 상단 배너
    _put(ws, 2, 2, f"- {y}년 {mth}월 -", font=F_BANNER, fill=NAVY, align=CENTER)
    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=ncol)
    _put(ws, 3, 2, "시선인터내셔널 Ad Report", font=F_BANNER_BIG, fill=NAVY, align=CENTER)
    ws.merge_cells(start_row=3, start_column=2, end_row=3, end_column=ncol)
    _put(ws, 5, 2, "[브랜드 별 매체 총 누적]", font=F_SEC)
    # 헤더 (row6): 브랜드 C유형 D:E매체 F집행율 G~ 지표
    _put(ws, 6, 2, "브랜드", font=F_COL, fill=FILL_COL, align=CENTER)
    _put(ws, 6, 3, "유형", font=F_COL, fill=FILL_COL, align=CENTER)
    _put(ws, 6, 4, "매체", font=F_COL, fill=FILL_COL, align=CENTER)
    ws.merge_cells(start_row=6, start_column=4, end_row=6, end_column=5)
    _put(ws, 6, 6, "집행율", font=F_COL, fill=FILL_COL, align=CENTER)
    for i, k in enumerate(BS_KEYS):
        _put(ws, 6, 7 + i, k, font=F_COL, fill=FILL_COL, align=CENTER)

    center_v = Alignment(horizontal="center", vertical="center")
    r = 7
    for b in ["MI", "EBM", "IT"]:
        dfb = uni[uni["브랜드"] == b]
        brand_start = r
        for gubun in ["성과형", "노출형"]:
            type_start = r
            for _, label, media, pat in [x for x in BS_MEDIA if x[0] == gubun]:
                sub = dfb.iloc[0:0] if media == "_NONE_" else _slice(dfb, media, pat)
                m = _metrics(sub)
                _put(ws, r, 4, label, align=LEFT)
                ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=5)
                # 집행율 = 집행예산 / 월예산(Total [매체 예산 집행율] 기준). 예산 없으면 '-'
                bud = BS_BUDGET.get((media, pat.lstrip("_")))
                if bud:
                    _put(ws, r, 6, m["집행예산"] / bud, "0.00%", align=CENTER)
                else:
                    _put(ws, r, 6, "-", align=CENTER)
                for i, k in enumerate(BS_KEYS):
                    _put(ws, r, 7 + i, m[k], BS_FMT[i])
                if m["집행예산"] == 0:          # 집행 0 → 행 숨김(발생 시 자동 표시)
                    ws.row_dimensions[r].hidden = True
                r += 1
            ws.merge_cells(start_row=type_start, start_column=3, end_row=r - 1, end_column=3)
            cell = ws.cell(type_start, 3, gubun); cell.alignment = center_v
        # 브랜드 세로병합
        ws.merge_cells(start_row=brand_start, start_column=2, end_row=r - 1, end_column=2)
        bc = ws.cell(brand_start, 2, b); bc.font = F_SUM; bc.alignment = center_v
        # 소계 (B~F 병합)
        mt = _metrics(dfb)
        _put(ws, r, 2, "소계", font=F_SUM, fill=FILL_SUM, align=CENTER)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)   # B:E 병합
        # F: 소계 집행율 = 브랜드 총집행 / 총예산
        _put(ws, r, 6, mt["집행예산"] / TOTAL_BUDGET, "0.00%",
             font=F_SUM, fill=FILL_SUM, align=CENTER)
        for i, k in enumerate(BS_KEYS):
            _put(ws, r, 7 + i, mt[k], BS_FMT[i], font=F_SUM, fill=FILL_SUM)
        r += 1               # 브랜드 사이 빈행 없음

    # 표 전체 테두리 (헤더6행 ~ 마지막행, B~마지막열)
    box = Border(*[Side("thin", "D9D9D9")] * 4)
    for rr in range(6, r):
        for cc in range(2, ncol + 1):
            ws.cell(rr, cc).border = box

    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 3


# 리포트 추가 요청: 매체 세부 16종 (6월F 순서). 모든 브랜드에 전 매체 행 표시.
SUBTYPES = [
    ("Google Pmax", "Google", "pmax"), ("Google Keyword", "Google", "cpc"),
    ("Naver Brand SA", "Naver SA", "bsa"), ("Naver Keyword SA", "Naver SA", "cpc"),
    ("Naver Shopping SA", "Naver SA", "shopping"), ("Naver Place SA", "Naver SA", "place"),
    ("Naver Ambassador", "Naver SA", "Ambassador"),
    ("Naver DA_Smart", "Naver", "smart"), ("Naver DA_ADVoost", "Naver", "advoost"),
    ("KAKAO Bizboard", "KKO", "biz"), ("KAKAO Native", "KKO", "ntv"),
    ("KAKAO Catalog", "KKO", "ca"),
    ("Criteo", "Criteo", ""), ("RTB House", "RTB", ""),
    ("Instagram_성과형", "Meta", "pf"), ("Instagram_노출형(br)", "Meta", "br"),
]


def write_report_request(ws, uni, y, mth):
    ndays = calendar.monthrange(y, mth)[1]
    _put(ws, 2, 2, "■ 광고비용 데일리 집행현황", font=F_TITLE)
    # (3행 공백) 헤더 (row 4): 브랜드, 매체, 날짜1..N, 월누계
    _put(ws, 4, 2, "브랜드", font=F_COL, fill=FILL_COL, align=CENTER)
    _put(ws, 4, 3, "매체", font=F_COL, fill=FILL_COL, align=CENTER)
    for d in range(1, ndays + 1):
        _put(ws, 4, 3 + d, date(y, mth, d), "mm-dd", font=F_COL, fill=FILL_COL, align=CENTER)
    _put(ws, 4, 4 + ndays, "월 누계", font=F_COL, fill=FILL_COL, align=CENTER)

    def cost_series(dfb, media, pat):
        d = dfb[dfb["매체"] == media]
        if pat:
            d = d[d["캠페인"].str.contains(pat, case=False, regex=False)]
        return d.groupby("날짜키")["광고비용"].sum().to_dict()

    r = 5
    center_v = Alignment(horizontal="center", vertical="center")
    box = Border(*[Side("thin", "D9D9D9")] * 4)   # 브랜드 셀 테두리(전역과 동일)
    for b in ["전체", "MI", "EBM", "IT"]:
        dfb = uni if b == "전체" else uni[uni["브랜드"] == b]
        start = r
        for label, media, pat in SUBTYPES:      # 모든 매체 행 표시(0이어도)
            series = cost_series(dfb, media, pat)
            _put(ws, r, 3, label, align=LEFT)
            total = 0.0
            for dd in range(1, ndays + 1):
                dk = date(y, mth, dd).strftime("%Y%m%d")
                v = series.get(dk, 0)
                total += v
                _put(ws, r, 3 + dd, v, "#,##0")
            _put(ws, r, 4 + ndays, total, "#,##0", font=F_SUM)
            if total == 0:                      # 집행 0 → 행 숨김(발생하면 자동 표시)
                ws.row_dimensions[r].hidden = True
            r += 1
        # 브랜드 셀 세로 병합 + 테두리
        ws.merge_cells(start_row=start, start_column=2, end_row=r - 1, end_column=2)
        for rr in range(start, r):
            ws.cell(rr, 2).border = box
        cell = ws.cell(start, 2, b)
        cell.font = F_SUM
        cell.alignment = center_v

    # 표 전체(헤더4행~데이터끝, B~월누계열) 모든 테두리
    last_row = r - 1
    last_col = 4 + ndays
    for rr in range(4, last_row + 1):
        for cc in range(2, last_col + 1):
            ws.cell(rr, cc).border = box

    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 18
    for c in range(4, 4 + ndays):
        ws.column_dimensions[ws.cell(row=1, column=c).column_letter].width = 9
    ws.column_dimensions[ws.cell(row=1, column=4 + ndays).column_letter].width = 13
