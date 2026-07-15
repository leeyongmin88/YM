# -*- coding: utf-8 -*-
"""데일리 리포트 자동화 - 설정값
wb-ss-da-dailyreport-v14 스킬을 로컬 Python 파이프라인으로 구현.
"""
from pathlib import Path

# --- 경로 ---
# config.py는 pipeline/ 안 → 상위 폴더가 프로젝트 루트. 어디에 두든 자동 인식(이식성).
YM_ROOT = Path(__file__).resolve().parent.parent
RAW_DIR = YM_ROOT / "Raw"
OUT_DIR = YM_ROOT / "output"

# --- 월예산 파일 (매월 사용자가 이 엑셀만 수정하면 집행율에 반영) ---
BUDGET_FILE = RAW_DIR / "예산.xlsx"


def _num(v):
    try:
        return float(v) if v not in (None, "") else 0.0
    except (TypeError, ValueError):
        return 0.0


def load_media_table():
    """Raw/예산.xlsx → [(구분, 매체라벨, 통합매체, 패턴, {MI,EBM,IT})]. 없으면 [].
    헤더: 구분 | 매체 | 통합매체 | 패턴 | MI | EBM | IT. (통합매체/패턴 열 있으면 매체목록도 파일이 정함)
    → 매체 추가: 이 파일에 한 줄 추가(구분·매체·통합매체·패턴·예산)하면 리포트에 반영."""
    if not BUDGET_FILE.exists():
        return []
    from openpyxl import load_workbook
    rows = list(load_workbook(BUDGET_FILE, data_only=True).active.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    ci = {n: (header.index(n) if n in header else None)
          for n in ("구분", "매체", "통합매체", "패턴", "MI", "EBM", "IT")}
    out = []
    for r in rows[1:]:
        if not r or ci["구분"] is None or r[ci["구분"]] is None:
            continue
        def g(name):
            i = ci[name]
            return r[i] if (i is not None and i < len(r)) else None
        media = g("통합매체")
        out.append((
            str(g("구분")).strip(), str(g("매체")).strip(),
            (str(media).strip() if media is not None else None),
            (str(g("패턴")).strip() if g("패턴") is not None else ""),
            {b: _num(g(b)) for b in ("MI", "EBM", "IT")},
        ))
    return out


def load_budgets():
    """{(구분, 매체라벨): {brand:budget}} — 정액(JEONGAEK) 예산 조회용."""
    return {(g, l): b for g, l, m, p, b in load_media_table()}

# --- 광고비 보정 계수 (스킬 공통원칙) ---
# Meta·Google·Criteo = /0.9*1.1 , RTB·KKO = *1.1 , Naver 계열 그대로
COST_COEF = {
    "Meta":     1.1 / 0.9,
    "Google":   1.1 / 0.9,
    "Criteo":   1.1 / 0.9,
    "RTB":      1.1,
    "KKO":      1.1,
    "Naver":    1.0,
    "Naver SA": 1.0,
}

# --- 통합 시트 컬럼 순서 (A~Q, 17열) ---
UNIFIED_COLS = [
    "날짜", "날짜키", "매체", "브랜드", "캠페인", "광고그룹", "광고(소재)",
    "광고비용", "노출수", "클릭수", "GA구매", "GA구매수익", "GA세션",
    "매핑상태", "매칭키", "회원가입수", "회원가입세션",
]

# 유효 브랜드 (캠페인명 토큰 검증용)
BRANDS = {"MI", "IT", "EBM"}

# --- Naver SA 정액(고정비) 6계열: 월예산 ---
# 일별 광고비 = 월예산 / 해당월 일수, 데이터 존재 기간에 적용.
# 캠페인명 : (브랜드, 매칭키, 기본 월예산). 실제값은 예산파일(SA 네이버 브랜드검색/엠버서더형) 우선.
_JEONGAEK_DEFAULT = {
    "NAV_MI_SA_pf_bsa_정액":         ("MI",  "정액_MI_bsa",  9_240_000),
    "NAV_MI_SA_pf_Ambassador_정액":  ("MI",  "정액_MI_amb",  2_530_000),
    "NAV_IT_SA_pf_bsa_정액":         ("IT",  "정액_IT_bsa",  17_710_000),
    "NAV_IT_SA_pf_Ambassador_정액":  ("IT",  "정액_IT_amb",  3_157_000),
    "NAV_EBM_SA_pf_bsa_정액":        ("EBM", "정액_EBM_bsa", 1_540_000),
    "NAV_EBM_SA_pf_Ambassador_정액": ("EBM", "정액_EBM_amb", 1_892_000),
}


# 정액 대상 매체: 예산파일 매체라벨(기기 접미사 제외) → 캠페인 종류
_JEONGAEK_KINDS = {"네이버 브랜드검색": "bsa", "네이버 엠버서더형": "Ambassador"}


def _split_device(label):
    """라벨 끝 ' PC'/' MO' → (base, 'pc'/'mo'/''). 기기 분리 예산 지원."""
    for suf, dev in ((" PC", "pc"), (" MO", "mo")):
        if label.endswith(suf):
            return label[:-len(suf)].rstrip(), dev
    return label, ""


def _build_jeongaek():
    """정액 월예산 {campaign:(brand,matchkey,budget)}: 예산파일에서 브랜드검색/엠버서더형 읽음.
    라벨에 ' PC'/' MO' 있으면 캠페인에 기기태그(_pc_/_mo_) 부여 → 일자별 PC/MO 분리.
    → 예산파일만 수정하면 정액 집행액·집행율 함께 반영. 파일 없으면 기본값."""
    out = {}
    for gubun, label, media, pat, budgets in load_media_table():
        base, dev = _split_device(label)
        kind = _JEONGAEK_KINDS.get(base)
        if not kind:
            continue
        for brand in ("MI", "EBM", "IT"):
            devtag = f"{dev}_" if dev else ""
            camp = f"NAV_{brand}_SA_pf_{kind}_{devtag}정액"
            key = f"정액_{brand}_{kind}"
            out[camp] = (brand, key, budgets.get(brand, 0))
    return out or dict(_JEONGAEK_DEFAULT)


JEONGAEK = _build_jeongaek()

# --- GA 조인 룩업 (추가정보 시트 기반, 매월 갱신 가능) ---
# KKO catalog: GA 세션캠페인(카탈로그) → KK코드
KKO_CATALOG = {
    "ebm_catalog": "KK0003",
    "it_catalog":  "KK0001",
    "mi_catalog":  "KK0002",
}
