# -*- coding: utf-8 -*-
"""전 시트 공통 디자인 마감: 글꼴 통일 + 얇은 테두리 + 정렬.

기존 채움색/폰트색(주말색 등)/집행현황 테두리는 보존하고 글꼴 패밀리만 통일.
"""
from openpyxl.styles import Font, Border, Side, Alignment

FONT = "맑은 고딕"
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
VCENTER = Alignment(vertical="center")

# 테두리 자동적용 제외 (자체 스타일 보유)
BORDER_SKIP = {"●광고비집행현황"}


def apply_global_style(book):
    for ws in book.worksheets:
        skip_border = ws.title in BORDER_SKIP
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                o = cell.font
                # 글꼴 패밀리만 통일, 볼드/크기/색 보존
                cell.font = Font(name=FONT, size=o.size or 10, bold=o.bold,
                                 italic=o.italic, color=o.color)
                # 정렬: 기존 수평정렬 보존 + 수직 가운데
                a = cell.alignment
                cell.alignment = Alignment(horizontal=a.horizontal,
                                           vertical="center", wrap_text=a.wrap_text)
                # 제목·섹션제목(12pt 이상)은 테두리 제외
                if not skip_border and (o.size or 10) < 12:
                    cell.border = BORDER
        if not skip_border:
            ws.sheet_view.showGridLines = False   # 테두리 있으니 격자선 숨김
        _autofit(ws)


def _disp_len(cell):
    """셀 표시폭(글자수) 추정 — 숫자서식/퍼센트/날짜 반영."""
    v = cell.value
    if v is None:
        return 0
    if isinstance(v, str):
        return 0 if v.startswith("=") else len(v)   # 수식은 제외
    fmt = cell.number_format or ""
    if isinstance(v, (int, float)):
        if "%" in fmt:
            return len(f"{v * 100:,.2f}%")
        if "0.00" in fmt:
            return len(f"{v:,.2f}")
        return len(f"{v:,.0f}")
    if hasattr(v, "strftime"):
        return 10                                   # yyyy-mm-dd
    return len(str(v))


def _autofit(ws):
    """열 너비 자동조정 (### 방지). 가로병합/제목(14pt)은 폭 계산 제외."""
    skip = set()
    for m in ws.merged_cells.ranges:
        if m.min_col != m.max_col:                  # 가로병합: 전 셀 제외
            for row in range(m.min_row, m.max_row + 1):
                for col in range(m.min_col, m.max_col + 1):
                    skip.add((row, col))
        else:                                       # 세로병합: 앵커 외 제외
            for row in range(m.min_row + 1, m.max_row + 1):
                skip.add((row, m.min_col))
    widths = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None or (cell.row, cell.column) in skip:
                continue
            if (cell.font.size or 10) >= 14:        # 제목 제외
                continue
            widths[cell.column] = max(widths.get(cell.column, 0), _disp_len(cell))
    for col, L in widths.items():
        letter = ws.cell(row=1, column=col).column_letter
        ws.column_dimensions[letter].width = min(max(L * 1.2 + 2, 6), 40)
