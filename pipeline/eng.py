# -*- coding: utf-8 -*-
"""통합_ENG 캠페인 시트.

메타 브랜딩형 중 광고그룹명에 '(eng)'가 포함된(게시물 참여 목표) 캠페인의
게시물 참여 지표를 브랜드 구분 없이 하나의 시트에 모아 보여준다.

소스: Raw/Meta_ENG/Meta_{MI,IT,EBM}_데일리_eng.xlsx (게시물 참여 전용 리포트).
  - 일반 Meta 폴더와 분리된 소스 → 통합(main) 리포트 로직에는 영향 없음.
  - 게시물 참여 지표 9개는 raw 컬럼 순서 그대로 재현(중복 라벨 포함).

레이아웃(참고파일 '시선닷컴_ENG 캠페인 시트.xlsx'):
  ① 상단: 캠페인(광고그룹)별 월 누적 요약 + 합계
  ② 하단: 일자 × 광고그룹 상세
"""
import warnings
warnings.simplefilter("ignore")
import pandas as pd
from openpyxl import load_workbook

from config import RAW_DIR, COST_COEF
from ingest import to_num, to_date
from total import (_put, excel_weeknum2, F_TITLE, F_COL, F_SUM,
                   FILL_COL, FILL_SUM, CENTER, LEFT, SAT_COLOR, SUN_COLOR)

ENG_DIR = RAW_DIR / "Meta_ENG"
BRAND_ORDER = ["MI", "EBM", "IT"]          # 표시 순서(데이터 없는 브랜드는 생략됨)

# 게시물 참여 지표 (raw 'Raw Data Report' 컬럼 순서 그대로 = 인덱스 8~16, 중복 라벨 포함)
ENG_METRIC_HDR = ["게시물 참여", "게시물참여(동영상재생 제외)", "게시물저장", "게시물 참여",
                  "게시물 댓글", "게시물 참여", "게시물 저장", "게시물 공감", "게시물 공유 수"]
N_ENG = len(ENG_METRIC_HDR)                # 9

# 시트 지표 컬럼 F~S (배송 5개 + 게시물 참여 9개)
METRIC_HDR = ["노출수", "클릭수", "클릭률", "클릭당비용", "집행예산"] + ENG_METRIC_HDR
METRIC_FMT = ["#,##0", "#,##0", "0.00%", "#,##0", "#,##0"] + ["#,##0"] * N_ENG
C_METRIC0 = 6                              # 지표 시작열 = F
_E_COLS = [f"e{i}" for i in range(N_ENG)]
_NUM_COLS = ["노출수", "클릭수", "지출_raw"] + _E_COLS


def _brand_from_file(name: str) -> str:
    up = name.upper()
    for b in ("MI", "IT", "EBM"):
        if f"_{b}_" in up:
            return b
    return ""


def load_eng() -> pd.DataFrame:
    """Raw/Meta_ENG 3파일 → (날짜, 브랜드, 광고그룹, 배송지표, 게시물참여9) 행.
    광고 단위 행을 (날짜×브랜드×광고그룹)로 합산하지 않고 원행 그대로 반환(집계는 호출부).
    '어제까지'만 반영(당일 이후 제외) — 통합 리포트와 동일 기준."""
    cols = ["날짜", "브랜드", "광고그룹"] + _NUM_COLS
    if not ENG_DIR.exists():
        return pd.DataFrame(columns=cols)
    recs = []
    for f in sorted(ENG_DIR.glob("*.xlsx")):
        brand = _brand_from_file(f.name)
        wb = load_workbook(f, read_only=False, data_only=True)
        ws = wb["Raw Data Report"] if "Raw Data Report" in wb.sheetnames else wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        for r in rows[1:]:
            if not r or r[0] is None:
                continue
            adgroup = str(r[2] or "")
            if "(eng)" not in adgroup.lower():
                continue
            eng = [to_num(r[8 + i]) for i in range(N_ENG)]
            recs.append([to_date(r[0]), brand, adgroup,
                         to_num(r[6]), to_num(r[7]), to_num(r[5]), *eng])
    df = pd.DataFrame(recs, columns=cols)
    if df.empty:
        return df
    df["날짜"] = pd.to_datetime(df["날짜"], errors="coerce")
    df = df[df["날짜"].notna()]
    today = pd.Timestamp("today").normalize()
    df = df[df["날짜"] < today].reset_index(drop=True)
    return df


def _metric_values(row) -> list:
    """합산 행(노출수/클릭수/지출_raw/e0..e8) → 시트 지표 14개 값.
    집행예산·클릭당비용은 통합 리포트와 동일하게 Meta 보정계수를 적용한다."""
    imp = row["노출수"]; clk = row["클릭수"]
    spend = row["지출_raw"] * COST_COEF["Meta"]
    ctr = clk / imp if imp else 0.0
    cpc = spend / clk if clk else 0.0
    eng = [row[c] for c in _E_COLS]
    return [imp, clk, ctr, cpc, spend] + eng


def _write_metric_headers(ws, r):
    for i, h in enumerate(METRIC_HDR):
        _put(ws, r, C_METRIC0 + i, h, font=F_COL, fill=FILL_COL, align=CENTER)


def _write_metrics(ws, r, values, font=None, fill=None):
    for i, v in enumerate(values):
        _put(ws, r, C_METRIC0 + i, v, METRIC_FMT[i], font=font, fill=fill)


def _bo(series):
    return series.map({b: i for i, b in enumerate(BRAND_ORDER)}).fillna(99)


def write_eng_sheet(ws, y, mth):
    df = load_eng()
    _put(ws, 2, 2, "통합_ENG 캠페인", font=F_TITLE)

    # ── ① 상단: 캠페인(광고그룹)별 누적 요약 ──
    hr = 4
    _put(ws, hr, 2, "브랜드", font=F_COL, fill=FILL_COL, align=CENTER)
    ws.merge_cells(start_row=hr, start_column=2, end_row=hr, end_column=4)
    _put(ws, hr, 5, "광고그룹", font=F_COL, fill=FILL_COL, align=CENTER)
    _write_metric_headers(ws, hr)
    r = hr + 1
    if not df.empty:
        summ = df.groupby(["브랜드", "광고그룹"], as_index=False)[_NUM_COLS].sum()
        summ = summ.assign(_bo=_bo(summ["브랜드"])).sort_values(["_bo", "광고그룹"])
        for _, row in summ.iterrows():
            _put(ws, r, 2, row["브랜드"], align=CENTER)
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
            _put(ws, r, 5, row["광고그룹"], align=LEFT)
            _write_metrics(ws, r, _metric_values(row))
            r += 1
        tot = df[_NUM_COLS].sum()
        _put(ws, r, 2, "합계", font=F_SUM, fill=FILL_SUM, align=CENTER)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
        _write_metrics(ws, r, _metric_values(tot), font=F_SUM, fill=FILL_SUM)
        r += 1
    else:
        _put(ws, r, 2, "(eng) 데이터 없음 — Raw/Meta_ENG 확인", align=LEFT)
        r += 1

    # ── ② 하단: 일자 × 광고그룹 상세 ──
    r += 1                                                  # 요약과 상세 사이 빈 행
    dh = r
    for c, h in enumerate(["주차", "날짜", "브랜드", "광고그룹"], start=2):
        _put(ws, dh, c, h, font=F_COL, fill=FILL_COL, align=CENTER)
    _write_metric_headers(ws, dh)
    r = dh + 1
    if not df.empty:
        daily = df.groupby(["날짜", "브랜드", "광고그룹"], as_index=False)[_NUM_COLS].sum()
        daily = daily.assign(_bo=_bo(daily["브랜드"])).sort_values(["날짜", "_bo", "광고그룹"])
        for _, row in daily.iterrows():
            d = row["날짜"]
            wd = d.weekday()
            col = SAT_COLOR if wd == 5 else SUN_COLOR if wd == 6 else None
            _put(ws, r, 2, excel_weeknum2(d.date()), align=CENTER)
            _put(ws, r, 3, d.to_pydatetime(), "yyyy-mm-dd", align=CENTER, color=col)
            _put(ws, r, 4, row["브랜드"], align=CENTER)
            _put(ws, r, 5, row["광고그룹"], align=LEFT)
            _write_metrics(ws, r, _metric_values(row))
            r += 1

    ws.column_dimensions["A"].width = 3
