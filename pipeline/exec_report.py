# -*- coding: utf-8 -*-
"""Phase 4: ●광고비집행현황 (다브랜드 일별 광고비 매트릭스).

일자 세로 × (시선전체/MI/EBM/IT) 블록. 각 블록: 전체·SA소계·DA소계 + 매체별.
값 = 광고비용 합계 (통합 기준).
"""
import warnings
warnings.simplefilter("ignore")
import calendar
from datetime import date
from openpyxl.styles import PatternFill, Border, Side, Font
from total import (week_in_month, WD_KR, _put, F_TITLE, SAT_COLOR, SUN_COLOR, CENTER)

# 6월 통합(F) 스타일
FILL_TITLE = PatternFill("solid", fgColor="222A35")   # 그룹제목 네이비
FILL_TOT = PatternFill("solid", fgColor="D8D8D8")     # 전체 회색
FILL_SA = PatternFill("solid", fgColor="FFFF00")      # SA 노랑
FILL_DA = PatternFill("solid", fgColor="E2EFD9")      # DA 연녹
FONT_TITLE = Font(name="맑은 고딕", bold=True, size=10, color="FFFFFF")
FONT_TOT = Font(name="맑은 고딕", bold=True, size=11, color="0000FF")
FONT_SUB = Font(name="맑은 고딕", bold=True, size=11)
FONT_HDR = Font(name="맑은 고딕", bold=True, size=10)
MED = Side(style="medium", color="000000")
HAIR = Side(style="hair", color="000000")
THIN = Side(style="thin", color="000000")

SA_ITEMS = ["N검색", "D검색", "G검색"]
# (그룹헤더, 소계라벨, 브랜드(None=전체), DA항목들)
BLOCKS = [
    ("광고비용-시선 전체", "시선 전체", None,
     ["K디스", "G디스", "N디스", "크리테오", "RTB하우스", "I디스"]),
    ("광고비용-미샤", "MI 소계", "MI",
     ["K디스", "G디스", "N디스", "크리테오", "크루비", "RTB하우스", "I디스"]),
    ("광고비용-E.B.M", "E.B.M 소계", "EBM",
     ["K디스", "G디스", "N디스", "크리테오", "버즈빌", "I디스"]),
    ("광고비용-잇미샤", "IT 소계", "IT",
     ["K디스", "G디스", "N디스", "크리테오", "RTB하우스", "모비온", "I디스", "틱톡"]),
]


def _cat(d, cat):
    """매체 카테고리별 통합 서브셋. 미보유 네트워크(D검색/크루비 등)는 빈 셋."""
    m, camp = d["매체"], d["캠페인"]
    has_cpc = camp.str.contains("cpc", case=False, regex=False)
    if cat == "N검색":   return d[m == "Naver SA"]
    if cat == "G검색":   return d[(m == "Google") & has_cpc]
    if cat == "G디스":   return d[(m == "Google") & ~has_cpc]
    if cat == "K디스":   return d[m == "KKO"]
    if cat == "N디스":   return d[m == "Naver"]
    if cat == "크리테오": return d[m == "Criteo"]
    if cat == "RTB하우스": return d[m == "RTB"]
    if cat == "I디스":   return d[m == "Meta"]
    return d.iloc[0:0]   # D검색·크루비·버즈빌·모비온·틱톡 = 0


def _fills(da_items, data=False):
    """컬럼별 채움. data=True면 상세(N/D/G검색·DA항목)는 무색, 소계·전체만 채움."""
    detail = None if data else FILL_SA
    detail_da = None if data else FILL_DA
    return ([FILL_TOT, FILL_SA] + [detail] * len(SA_ITEMS)
            + [FILL_DA] + [detail_da] * len(da_items))


def _fonts(da_items):
    """전체=파랑볼드, 소계=볼드, 상세=기본."""
    return [FONT_TOT, FONT_SUB] + [None] * len(SA_ITEMS) + [FONT_SUB] + [None] * len(da_items)


def write_exec_report(ws, uni, y, mth):
    ndays = calendar.monthrange(y, mth)[1]
    last_row = 7 + ndays
    ws.sheet_view.showGridLines = False
    _put(ws, 2, 2, "■ 광고비용 집행현황", font=F_TITLE)
    # 구분 (B4:D5 병합) + 일자/요일/주차 (row6)
    ws.merge_cells("B4:D5")
    _put(ws, 4, 2, "구분", font=FONT_TITLE, fill=FILL_TITLE, align=CENTER)
    for cc, h in [(2, "일자"), (3, "요일"), (4, "주차")]:
        _put(ws, 6, cc, h, font=FONT_TITLE, fill=FILL_TITLE, align=CENTER)

    blocks = []
    c = 5   # A열 비움 → B부터, 데이터블록은 E부터
    nsa = len(SA_ITEMS)
    for gtitle, subtitle, brand, da_items in BLOCKS:
        dfb = uni if brand is None else uni[uni["브랜드"] == brand]
        item_daily = {it: _cat(dfb, it).groupby("날짜키")["광고비용"].sum().to_dict()
                      for it in SA_ITEMS + da_items}
        cols = ["전체", "소계"] + SA_ITEMS + ["소계"] + da_items
        ncols = len(cols)
        hfills = _fills(da_items, data=False)
        dfills = _fills(da_items, data=True)
        fonts = _fonts(da_items)
        # 그룹제목 (row4, 블록 전체 병합, 네이비)
        ws.merge_cells(start_row=4, start_column=c, end_row=4, end_column=c + ncols - 1)
        _put(ws, 4, c, gtitle, font=FONT_TITLE, fill=FILL_TITLE, align=CENTER)
        # row5: 전체(subtitle) + SA병합(노랑) + DA병합(연녹)
        _put(ws, 5, c, subtitle, font=FONT_TOT, fill=FILL_TOT, align=CENTER)
        ws.merge_cells(start_row=5, start_column=c + 1, end_row=5, end_column=c + 1 + nsa)
        _put(ws, 5, c + 1, "SA", font=FONT_HDR, fill=FILL_SA, align=CENTER)
        ws.merge_cells(start_row=5, start_column=c + 2 + nsa, end_row=5, end_column=c + ncols - 1)
        _put(ws, 5, c + 2 + nsa, "DA", font=FONT_HDR, fill=FILL_DA, align=CENTER)
        # row6 컬럼헤더
        for i, cn in enumerate(cols):
            _put(ws, 6, c + i, cn, font=FONT_TOT if i == 0 else FONT_HDR,
                 fill=hfills[i], align=CENTER)
        blocks.append((c, ncols, da_items, item_daily, dfills, fonts))
        c += ncols
    end_col = c - 1

    def write_vals(row, c0, da_items, item_daily, dfills, fonts, dk, bold=False):
        get = (lambda it: sum(item_daily[it].values())) if dk is None \
            else (lambda it: item_daily[it].get(dk, 0))
        sa = sum(get(it) for it in SA_ITEMS)
        da = sum(get(it) for it in da_items)
        vals = [sa + da, sa] + [get(it) for it in SA_ITEMS] + [da] + [get(it) for it in da_items]
        for i, v in enumerate(vals):
            fnt = fonts[i] or (FONT_HDR if bold else None)
            _put(ws, row, c0 + i, v, "#,##0", font=fnt, fill=dfills[i])

    # 누적 (row 7)
    _put(ws, 7, 2, "누적", font=FONT_TOT, fill=FILL_TOT, align=CENTER)
    _put(ws, 7, 3, None, fill=FILL_TOT)
    _put(ws, 7, 4, None, fill=FILL_TOT)
    for c0, ncols, da_items, item_daily, dfills, fonts in blocks:
        write_vals(7, c0, da_items, item_daily, dfills, fonts, None, bold=True)
    # 일자별 (row 8~)
    for day in range(1, ndays + 1):
        d = date(y, mth, day); dk = d.strftime("%Y%m%d"); row = 7 + day
        col = SAT_COLOR if d.weekday() == 5 else SUN_COLOR if d.weekday() == 6 else None
        _put(ws, row, 2, d, "yyyy-mm-dd", align=CENTER, color=col)
        _put(ws, row, 3, WD_KR[d.weekday()], align=CENTER, color=col)
        _put(ws, row, 4, week_in_month(d), align=CENTER)
        for c0, ncols, da_items, item_daily, dfills, fonts in blocks:
            write_vals(row, c0, da_items, item_daily, dfills, fonts, dk)

    # 테두리: 블록경계 medium, 내부 hair, 행 thin
    starts = {2} | {b[0] for b in blocks}
    ends = {4} | {b[0] + b[1] - 1 for b in blocks}
    for row in range(4, last_row + 1):
        for col in range(2, end_col + 1):
            left = MED if col in starts else HAIR
            right = MED if col in ends else HAIR
            ws.cell(row=row, column=col).border = Border(left=left, right=right, top=THIN, bottom=THIN)

    # 열 너비 (A 비움)
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 5
    ws.column_dimensions["D"].width = 5
    for cc in range(5, end_col + 1):
        ws.column_dimensions[ws.cell(row=1, column=cc).column_letter].width = 11
