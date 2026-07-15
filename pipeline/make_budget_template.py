# -*- coding: utf-8 -*-
"""Raw/예산.xlsx 템플릿 생성 (매체×브랜드 월예산). 현재 코드값으로 채움."""
import os
import sys
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from total import MEDIA_ROWS
from config import JEONGAEK, BUDGET_FILE
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# JEONGAEK: (brand, 'bsa'/'amb') → 월예산 (브랜드별 정액값)
jb = {}
for camp, (brand, key, bud) in JEONGAEK.items():
    jb[(brand, "bsa" if "bsa" in camp else "amb")] = bud


def brand_vals(label, default):
    lab = label.strip()
    if lab == "네이버 브랜드검색":
        return {b: jb.get((b, "bsa"), default) for b in ("MI", "EBM", "IT")}
    if lab == "네이버 엠버서더형":
        return {b: jb.get((b, "amb"), default) for b in ("MI", "EBM", "IT")}
    return {b: default for b in ("MI", "EBM", "IT")}


wb = Workbook(); ws = wb.active; ws.title = "예산"
navy = PatternFill("solid", fgColor="FF1F4E78")
hfont = Font(name="맑은 고딕", bold=True, color="FFFFFFFF")
side = Side(style="thin", color="FFBFBFBF")
box = Border(left=side, right=side, top=side, bottom=side)
ctr = Alignment(horizontal="center", vertical="center")

for c, h in enumerate(["구분", "매체", "통합매체", "패턴", "MI", "EBM", "IT"], 1):
    cell = ws.cell(1, c, h); cell.font = hfont; cell.fill = navy; cell.alignment = ctr

r = 2
for gubun, label, media, pat, default in MEDIA_ROWS:
    v = brand_vals(label, default)
    ws.cell(r, 1, gubun); ws.cell(r, 2, label.strip())
    ws.cell(r, 3, media); ws.cell(r, 4, pat)
    ws.cell(r, 5, v["MI"]); ws.cell(r, 6, v["EBM"]); ws.cell(r, 7, v["IT"])
    for c in (5, 6, 7):
        ws.cell(r, c).number_format = "#,##0"
    r += 1

for rr in range(1, r):
    for cc in range(1, 8):
        ws.cell(rr, cc).border = box
for L, wd in [("A", 12), ("B", 18), ("C", 12), ("D", 12), ("E", 13), ("F", 13), ("G", 13)]:
    ws.column_dimensions[L].width = wd
ws.sheet_view.showGridLines = False

wb.save(BUDGET_FILE)
print("생성 완료:", BUDGET_FILE)
