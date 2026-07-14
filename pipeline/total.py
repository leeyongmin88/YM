# -*- coding: utf-8 -*-
"""Phase 2: 브랜드 Total 대시보드 (MI/IT/EBM).

통합에서 브랜드별 슬라이스 → 6개 섹션 값 산출 후 시트 작성:
 [매체 예산 집행율] [매체 총 누적] [디바이스별=구분롤업] [요일별 평균] [주간] [일자별 성과]
"""
import warnings
warnings.simplefilter("ignore")
import calendar
from datetime import date, timedelta
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from config import load_budgets

# 월예산 파일(Raw/예산.xlsx) 로드. {(구분,라벨):{brand:budget}}. 없으면 {} → MEDIA_ROWS 기본값.
BUDGETS = load_budgets()


def budget_of(brand, gubun, label, default=0):
    """브랜드별 월예산: 예산파일 우선, 없으면 코드 기본값(default)."""
    return BUDGETS.get((str(gubun).strip(), str(label).strip()), {}).get(brand, default)

BRAND_TITLE = {"MI": "미샤", "IT": "잇미샤", "EBM": "E.B.M"}

# 스타일
# 색상은 모두 불투명 8자리(FFxxxxxx): 알파 00이면 엑셀서 투명하게 렌더됨
F_TITLE = Font(bold=True, size=14)
F_SEC = Font(bold=True, size=12, color="FF1F4E78")    # 섹션 제목: 파란 글씨(채움없음)
F_COL = Font(bold=True, size=10, color="FFFFFFFF")    # 컬럼헤더: 흰 글씨
F_SUM = Font(bold=True)
FILL_SEC = None                                        # 섹션 제목 채움 없음
FILL_COL = PatternFill("solid", fgColor="FF1F4E78")   # 컬럼헤더 네이비
FILL_SUM = PatternFill("solid", fgColor="FFF2F2F2")   # 합계 음영(BFBFBF 테두리 보이게 밝은 회색)
SAT_COLOR = "FF0000FF"                                  # 토 파랑
SUN_COLOR = "FFFF0000"                                  # 일 빨강
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")

# 숫자서식: 노출,클릭,클릭률,클릭당비용,집행예산,전환수,매출,회원가입,전환율,전환당비용,회원가입율,ROAS,객단가
# MI_Total_F 순서: ...매출, 세션수, 회원가입, 전환율, 전환당비용, 세션당비용, 회원가입율, ROAS, 객단가
CUM_KEYS = ["노출수", "클릭수", "클릭률", "클릭당비용", "집행예산", "전환수", "매출",
            "세션수", "회원가입", "전환율", "전환당비용", "세션당비용",
            "회원가입율", "ROAS", "객단가"]
CUM_FMT = ["#,##0", "#,##0", "0.00%", "#,##0", "#,##0", "#,##0", "#,##0",
           "#,##0", "#,##0", "0.00%", "#,##0", "#,##0",
           "0.00%", "#,##0.00", "#,##0"]

# (구분, 라벨, 매체, 캠페인 부분문자열, 월예산)   부분문자열 ""=매체 전체
MEDIA_ROWS = [
    ("SA",        "네이버 브랜드검색", "Naver SA", "bsa",        11_330_000),
    ("SA",        "네이버 키워드검색", "Naver SA", "cpc",         3_137_000),
    ("SA",        "네이버 쇼핑검색",   "Naver SA", "shopping",    2_000_000),
    ("SA",        "네이버 엠버서더형", "Naver SA", "Ambassador",  2_442_000),
    ("SA",        "구글 키워드검색",   "Google",   "cpc",         1_500_000),
    ("DA(성과형)", "카카오 네이티브",   "KKO",      "ntv",           940_000),
    ("DA(성과형)", "카카오 비즈보드",   "KKO",      "biz",           600_000),
    ("DA(성과형)", "카카오 카탈로그",   "KKO",      "_ca",         3_000_000),
    ("DA(성과형)", "구글 쇼핑",        "Google",   "pmax",        9_191_000),
    ("DA(성과형)", "크리테오",         "Criteo",   "",           17_000_000),
    ("DA(성과형)", "인스타그램 ",      "Meta",     "pf",          8_160_000),
    ("DA(성과형)", "RTB",             "RTB",      "",            2_700_000),
    ("DA(성과형)", "신규 매체",        "_NONE_",   "",            3_000_000),
    ("DA(성과형)", "구글 GDN",         "Google",   "gdn",                 0),
    ("DA(노출형)", "인스타그램 ",      "Meta",     "br",          2_200_000),
    ("DA(노출형)", "구글 YouTube",     "Google",   "youtube",             0),
]

CUM_HDR = ["노출수", "클릭수", "클릭률", "클릭당비용", "집행예산", "전환수", "매출",
           "회원가입", "전환율", "전환당비용", "회원가입율", "ROAS", "객단가"]
WD_KR = ["월", "화", "수", "목", "금", "토", "일"]

# 세션수·세션당단가 = 세션 관련 지표 (2026-07 추가)


def _div(a, b):
    return a / b if b else 0.0


def _slice(df, media, pattern):
    m = df["매체"] == media
    if pattern:
        m &= df["캠페인"].str.contains(pattern, case=False, regex=False)
    return df[m]


def _build_metrics(imp, clk, cost, cv, rev, mem, sess):
    return {
        "노출수": imp, "클릭수": clk, "클릭률": _div(clk, imp), "클릭당비용": _div(cost, clk),
        "집행예산": cost, "전환수": cv, "매출": rev, "회원가입": mem, "세션수": sess,
        "전환율": _div(cv, clk), "전환당비용": _div(cost, cv), "회원가입율": _div(mem, clk),
        "세션당비용": _div(cost, sess), "ROAS": _div(rev, cost), "객단가": _div(rev, cv),
    }


def _metrics(sub):
    return _build_metrics(sub["노출수"].sum(), sub["클릭수"].sum(), sub["광고비용"].sum(),
                          sub["GA구매"].sum(), sub["GA구매수익"].sum(),
                          sub["회원가입수"].sum(), sub["GA세션"].sum())


def _metrics_from_sums(imp, clk, cost, cv, rev, mem, sess=0):
    return _build_metrics(imp, clk, cost, cv, rev, mem, sess)


def excel_weeknum2(d):
    jan1 = date(d.year, 1, 1)
    start = jan1 - timedelta(days=jan1.weekday())     # Jan1 주의 월요일
    return (d - start).days // 7 + 1


def week_in_month(d):
    first = d.replace(day=1)
    return excel_weeknum2(d) - excel_weeknum2(first) + 1


def media_cumulative(df_brand, brand):
    """[매체 총 누적] 각 라벨 지표 리스트 + 합계. 예산=브랜드별 파일값 우선(없으면 기본)."""
    rows = []
    for gubun, label, media, pat, default in MEDIA_ROWS:
        bud = budget_of(brand, gubun, label, default)
        rows.append((gubun, label, bud, _metrics(_slice(df_brand, media, pat))))
    return rows, _metrics(df_brand)


def gubun_rollup(cum_rows):
    """구분(SA/DA성과형/DA노출형)별 지표 롤업."""
    order = ["SA", "DA(성과형)", "DA(노출형)"]
    acc = {g: dict(노출수=0, 클릭수=0, 집행예산=0, 전환수=0, 매출=0, 회원가입=0, 세션수=0) for g in order}
    for gubun, label, budget, m in cum_rows:
        a = acc[gubun]
        for k in a:
            a[k] += m[k]
    out = []
    for g in order:
        a = acc[g]
        out.append((g, _metrics_from_sums(a["노출수"], a["클릭수"], a["집행예산"],
                                          a["전환수"], a["매출"], a["회원가입"], a["세션수"])))
    return out


def weekday_avg(daily_df, y, mth):
    """요일별 평균 = 해당 요일 합계 / 그 달 해당 요일 수. 주중/주말/일 평균 포함."""
    ndays = calendar.monthrange(y, mth)[1]
    wd_count = [0] * 7
    for day in range(1, ndays + 1):
        wd_count[date(y, mth, day).weekday()] += 1
    rows = []
    for wd in range(7):
        sub = daily_df[daily_df["wd"] == wd]
        c = wd_count[wd] or 1
        imp = sub["노출수"].sum() / c; clk = sub["클릭수"].sum() / c
        cost = sub["집행예산"].sum() / c; cv = sub["전환수"].sum() / c
        rev = sub["매출"].sum() / c
        mem = sub["회원가입"].sum() / c; sess = sub["세션수"].sum() / c
        rows.append({"요일": WD_KR[wd], "wd": wd, "노출": imp, "클릭": clk,
                     "클릭율": _div(clk, imp), "클릭비용": _div(cost, clk), "광고비": cost,
                     "전환": cv, "전환비용": _div(cost, cv), "매출": rev,
                     "ROAS": _div(rev, cost), "객단가": _div(rev, cv),
                     "회원가입": mem, "세션수": sess, "전환율": _div(cv, clk),
                     "회원가입율": _div(mem, clk), "세션당비용": _div(cost, sess)})
    def avg(sel):
        keys = ["노출", "클릭", "광고비", "전환", "매출"]
        n = len(sel) or 1
        s = {k: sum(r[k] for r in sel) / n for k in keys}
        return {**s, "클릭율": _div(s["클릭"], s["노출"]), "클릭비용": _div(s["광고비"], s["클릭"]),
                "전환비용": _div(s["광고비"], s["전환"]), "ROAS": _div(s["매출"], s["광고비"]),
                "객단가": _div(s["매출"], s["전환"])}
    weekday = avg(rows[0:5])       # 월~금
    weekend = avg(rows[5:7])       # 토,일
    allavg = avg(rows)             # 전체
    return rows, weekday, weekend, allavg


def daily_frame(df_brand, y, mth):
    """1일~월말 전체 일자 성과 (빠진 날은 0)."""
    ndays = calendar.monthrange(y, mth)[1]
    g = df_brand.groupby("날짜키").agg(
        imp=("노출수", "sum"), clk=("클릭수", "sum"), cost=("광고비용", "sum"),
        cv=("GA구매", "sum"), rev=("GA구매수익", "sum"), mem=("회원가입수", "sum"),
        sess=("GA세션", "sum")).to_dict("index")
    recs = []
    for day in range(1, ndays + 1):
        d = date(y, mth, day)
        dk = d.strftime("%Y%m%d")
        s = g.get(dk, {"imp": 0, "clk": 0, "cost": 0, "cv": 0, "rev": 0, "mem": 0, "sess": 0})
        m = _metrics_from_sums(s["imp"], s["clk"], s["cost"], s["cv"], s["rev"], s["mem"], s["sess"])
        recs.append({"주차": week_in_month(d), "요일": WD_KR[d.weekday()], "날짜": d,
                     "wd": d.weekday(), **m})
    return pd.DataFrame(recs)


def _put(ws, r, c, v, fmt=None, font=None, fill=None, align=None, color=None):
    cell = ws.cell(row=r, column=c, value=v)
    if fmt:
        cell.number_format = fmt
    if font:
        cell.font = font
    elif color:
        cell.font = Font(color=color)
    if fill:
        cell.fill = fill
    if align:
        cell.alignment = align
    return cell


_TB_SIDE = Side(style="thin", color="FFBFBFBF")   # 중간 회색(style.py와 동일)
TOTAL_BORDER = Border(left=_TB_SIDE, right=_TB_SIDE, top=_TB_SIDE, bottom=_TB_SIDE)
def _merge_bc(ws, r, fill=None, font=None):
    """B:C 가로 병합 + 가운데. 필요 시 C에 색/폰트 적용해 밴드 연속."""
    _put(ws, r, 3, "", font=font, fill=fill)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
    ws.cell(row=r, column=2).alignment = CENTER


def write_total_sheet(ws, brand, df_brand, y, mth):
    """Total 대시보드 6개 섹션 작성. B열부터 시작."""
    cum_rows, total = media_cumulative(df_brand, brand)
    daily = daily_frame(df_brand, y, mth)

    _put(ws, 2, 2, f"- {y}년 {mth}월 -", font=F_TITLE)
    _put(ws, 3, 2, f"[ {BRAND_TITLE[brand]} ] Ad Report", font=F_TITLE)
    r = 6

    # ── 1. 매체 예산 집행율 ──
    _put(ws, r, 2, "[매체 예산 집행율]", font=F_SEC, fill=FILL_SEC)
    r += 1
    for c, h in enumerate(["구분", "매체별 성과", "월예산", "집행예산", "집행율", "비고"], start=2):
        _put(ws, r, c, h, font=F_COL, fill=FILL_COL, align=CENTER)
    r += 1
    tb = tc = 0.0
    for gubun, label, budget, m in cum_rows:
        _put(ws, r, 2, gubun, align=CENTER)
        _put(ws, r, 3, label, align=LEFT)
        _put(ws, r, 4, budget, "#,##0")
        _put(ws, r, 5, m["집행예산"], "#,##0")
        _put(ws, r, 6, _div(m["집행예산"], budget), "0.00%")
        _put(ws, r, 7, "")                        # 비고 열(테두리용 빈칸)
        if m["집행예산"] == 0:                      # 집행 0 매체행 숨김(집행 발생 시 자동 해제)
            ws.row_dimensions[r].hidden = True
        tb += budget; tc += m["집행예산"]
        r += 1
    # 합계: B:C 병합·가운데, 비고(G)도 합계색
    _put(ws, r, 2, "합계", font=F_SUM, fill=FILL_SUM, align=CENTER)
    _put(ws, r, 3, "", font=F_SUM, fill=FILL_SUM)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
    _put(ws, r, 4, tb, "#,##0", font=F_SUM, fill=FILL_SUM)
    _put(ws, r, 5, tc, "#,##0", font=F_SUM, fill=FILL_SUM)
    _put(ws, r, 6, _div(tc, tb), "0.00%", font=F_SUM, fill=FILL_SUM)
    _put(ws, r, 7, "", font=F_SUM, fill=FILL_SUM)
    r += 3

    # ── 2. 매체 총 누적 ──
    _put(ws, r, 2, "[매체 총 누적]", font=F_SEC, fill=FILL_SEC)
    r += 1
    _put(ws, r, 2, "구분", font=F_COL, fill=FILL_COL, align=CENTER)
    _put(ws, r, 3, "매체별 성과", font=F_COL, fill=FILL_COL, align=CENTER)
    for i, h in enumerate(CUM_KEYS):
        _put(ws, r, 4 + i, h, font=F_COL, fill=FILL_COL, align=CENTER)
    r += 1
    for gubun, label, budget, m in cum_rows:
        _put(ws, r, 2, gubun, align=CENTER)
        _put(ws, r, 3, label, align=LEFT)
        for i, k in enumerate(CUM_KEYS):
            _put(ws, r, 4 + i, m[k], CUM_FMT[i])
        if m["집행예산"] == 0:                      # 집행 0 매체행 숨김(집행 발생 시 자동 해제)
            ws.row_dimensions[r].hidden = True
        r += 1
    _put(ws, r, 2, "합계", font=F_SUM, fill=FILL_SUM, align=CENTER)
    _put(ws, r, 3, "", font=F_SUM, fill=FILL_SUM)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
    for i, k in enumerate(CUM_KEYS):
        _put(ws, r, 4 + i, total[k], CUM_FMT[i], font=F_SUM, fill=FILL_SUM)
    r += 3

    # ── 3. 광고 형태별 (구분 롤업) ──
    _put(ws, r, 2, "[광고 형태별]", font=F_SEC, fill=FILL_SEC)
    r += 1
    _put(ws, r, 2, "구분", font=F_COL, fill=FILL_COL, align=CENTER)
    _merge_bc(ws, r, fill=FILL_COL, font=F_COL)
    for i, h in enumerate(CUM_KEYS):
        _put(ws, r, 4 + i, h, font=F_COL, fill=FILL_COL, align=CENTER)
    r += 1
    for g, m in gubun_rollup(cum_rows):
        _put(ws, r, 2, g, align=CENTER)
        _merge_bc(ws, r)
        for i, k in enumerate(CUM_KEYS):
            _put(ws, r, 4 + i, m[k], CUM_FMT[i])
        r += 1
    r += 2

    # ── 4. 요일별 평균 (구분: 주중/주말) ──
    _put(ws, r, 2, "[요일별 평균] 통계 기간내", font=F_SEC, fill=FILL_SEC)
    r += 1
    r += 1                                          # (구 그룹헤더행) 비움: 텍스트·색·테두리 없음
    wa_hdr = ["평균 노출수", "평균 클릭수", "클릭율", "클릭비용", "평균광고비",
              "평균 전환수", "전환비용", "평균 매출", "ROAS", "객단가"]
    wa_key = ["노출", "클릭", "클릭율", "클릭비용", "광고비", "전환", "전환비용", "매출", "ROAS", "객단가"]
    wa_fmt = ["#,##0", "#,##0", "0.00%", "#,##0", "#,##0", "#,##0.0", "#,##0", "#,##0", "#,##0.00", "#,##0"]
    _put(ws, r, 2, "구분", font=F_COL, fill=FILL_COL, align=CENTER)   # B57 구분(C57과 동일서식)
    _put(ws, r, 3, "요일", font=F_COL, fill=FILL_COL, align=CENTER)
    for i, h in enumerate(wa_hdr):
        _put(ws, r, 4 + i, h, font=F_COL, fill=FILL_COL, align=CENTER)
    r += 1
    wrows, wmid, wend, wall = weekday_avg(daily, y, mth)
    groups = [("주중", wrows[0:5]), ("주말", wrows[5:7])]
    for gname, grp in groups:
        gstart = r
        for idx, wr in enumerate(grp):
            col = SAT_COLOR if wr["wd"] == 5 else SUN_COLOR if wr["wd"] == 6 else None
            if idx == 0:
                _put(ws, r, 2, gname, align=CENTER)
            _put(ws, r, 3, wr["요일"], align=CENTER, color=col)
            for i, k in enumerate(wa_key):
                _put(ws, r, 4 + i, wr[k], wa_fmt[i])
            r += 1
        # 주중(월~금)/주말(토·일) B열 세로 병합 + 가운데 + 테두리
        if r - 1 > gstart:
            ws.merge_cells(start_row=gstart, start_column=2, end_row=r - 1, end_column=2)
        ws.cell(row=gstart, column=2).alignment = CENTER
        for rr2 in range(gstart, r):                     # 병합 셀 전체 테두리
            ws.cell(row=rr2, column=2).border = TOTAL_BORDER
    for nm, mm in [("주중 평균", wmid), ("주말 평균", wend), ("일 평균", wall)]:
        _put(ws, r, 2, nm, font=F_SUM, fill=FILL_SUM, align=CENTER)
        _put(ws, r, 3, "", font=F_SUM, fill=FILL_SUM)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
        for i, k in enumerate(wa_key):
            _put(ws, r, 4 + i, mm[k], wa_fmt[i], font=F_SUM, fill=FILL_SUM)
        r += 1
    # (표 테두리는 apply_global_style이 전 표 동일하게 D9D9D9로 통일 적용)
    r += 2

    # ── 5. 주간 (주차번호 A열, CAC, 전주대비증감율) ──
    _put(ws, r, 2, "[주간]", font=F_SEC, fill=FILL_SEC)
    r += 1
    wk_hdr = ["노출수", "클릭수", "클릭률", "클릭당비용", "광고비", "전환수", "매출",
              "세션수", "회원가입", "전환율", "전환당비용", "세션당비용",
              "회원가입율", "ROAS", "객단가"]
    _put(ws, r, 2, "주간", font=F_COL, fill=FILL_COL, align=CENTER)
    _merge_bc(ws, r, fill=FILL_COL, font=F_COL)
    for i, h in enumerate(wk_hdr):
        _put(ws, r, 4 + i, h, font=F_COL, fill=FILL_COL, align=CENTER)
    _put(ws, r, 4 + len(wk_hdr), "CAC", font=F_COL, fill=FILL_COL, align=CENTER)
    r += 1
    wk_metrics = []
    for wk in range(1, 6):
        sub = daily[daily["주차"] == wk]
        m = _metrics_from_sums(sub["노출수"].sum(), sub["클릭수"].sum(), sub["집행예산"].sum(),
                               sub["전환수"].sum(), sub["매출"].sum(), sub["회원가입"].sum(),
                               sub["세션수"].sum())
        wk_metrics.append(m)
        _put(ws, r, 1, wk, align=CENTER, color="FFFFFFFF")    # A열 주차번호(흰글씨)
        _put(ws, r, 2, f"{mth}월 {wk}주", align=CENTER)
        _merge_bc(ws, r)
        for i, k in enumerate(CUM_KEYS):
            _put(ws, r, 4 + i, m[k], CUM_FMT[i])
        _put(ws, r, 4 + len(CUM_KEYS), _div(m["집행예산"], m["회원가입"]), "#,##0")  # CAC
        r += 1
    # 전주 대비 증감율: 데이터 있는 마지막 두 주 비교
    active = [i for i, m in enumerate(wk_metrics) if m["집행예산"] > 0]
    _put(ws, r, 2, "전주 대비 증감율", font=F_SUM, fill=FILL_SUM)
    _merge_bc(ws, r, fill=FILL_SUM, font=F_SUM)
    if len(active) >= 2:
        cur, prev = wk_metrics[active[-1]], wk_metrics[active[-2]]
        for i, k in enumerate(CUM_KEYS):
            v = _div(cur[k] - prev[k], prev[k]) if prev[k] else -1
            _put(ws, r, 4 + i, v, "0.00%", font=F_SUM, fill=FILL_SUM)
    else:
        for i in range(len(CUM_KEYS)):
            _put(ws, r, 4 + i, -1, "0.00%", font=F_SUM, fill=FILL_SUM)
    r += 3

    # ── 6. 일자별 성과 (주차번호 A열, 1일~월말, 주말색) ──
    _put(ws, r, 2, "[일자별 성과]", font=F_SEC, fill=FILL_SEC)
    _put(ws, r, 8, "어드민 전환")
    r += 1
    _put(ws, r, 2, "요일", font=F_COL, fill=FILL_COL, align=CENTER)
    _put(ws, r, 3, "날짜", font=F_COL, fill=FILL_COL, align=CENTER)
    for i, h in enumerate(wk_hdr):
        _put(ws, r, 4 + i, h, font=F_COL, fill=FILL_COL, align=CENTER)
    r += 1
    for _, d in daily.iterrows():
        col = SAT_COLOR if d["wd"] == 5 else SUN_COLOR if d["wd"] == 6 else None
        _put(ws, r, 1, int(d["주차"]), align=CENTER, color="FFFFFFFF")   # A열 주차번호(흰글씨)
        _put(ws, r, 2, d["요일"], align=CENTER, color=col)
        _put(ws, r, 3, d["날짜"], "yyyy-mm-dd", align=CENTER, color=col)
        for i, k in enumerate(CUM_KEYS):
            _put(ws, r, 4 + i, d[k], CUM_FMT[i])
        r += 1
    _put(ws, r, 2, "합계", font=F_SUM, fill=FILL_SUM, align=CENTER)
    _merge_bc(ws, r, fill=FILL_SUM, font=F_SUM)
    for i, k in enumerate(CUM_KEYS):
        _put(ws, r, 4 + i, total[k], CUM_FMT[i], font=F_SUM, fill=FILL_SUM)

    # 열 폭
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 16
    for c in range(4, 4 + len(CUM_KEYS)):
        ws.column_dimensions[ws.cell(row=1, column=c).column_letter].width = 13


if __name__ == "__main__":
    from build import build_unified
    uni = build_unified()
    for b in ["MI", "IT", "EBM"]:
        sub = uni[uni["브랜드"] == b]
        tot = _metrics(sub)
        print(f"[{b}] 광고비={tot['집행예산']:,.0f} 전환={tot['전환수']:.0f} 매출={tot['매출']:,.0f}")
